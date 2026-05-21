#!/usr/bin/env python3
"""
ZKTeco eFace10 Utility v4.3 — CV RAJ
Split panel: kiri workflow, kanan viewer + history
Semua data disimpan di SQLite, tidak ada file temp eksternal
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv, os, threading, calendar, sqlite3, json, sys
from datetime import datetime, date, timedelta

APP_VERSION = "4.3.0"
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
DB_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "absensi.db")

BULAN_ID = ["Januari","Februari","Maret","April","Mei","Juni",
            "Juli","Agustus","September","Oktober","November","Desember"]
HARI_ID  = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]

DEFAULT_CONFIG = {
    "ip": "10.10.11.55", "port": "8088",
    "lang": "en",
    "jam_masuk": "08:00", "jam_keluar": "16:00",
    "toleransi": 15, "auto_backup": False,
    "user_map": {
        "1":"NICHOLAS","2":"SERLI","3":"TIA","4":"MISRO",
        "5":"LISA","6":"TUR","7":"SLAMET","8":"ARI",
        "9":"REFA","10":"SUKUR","11":"PUGUH"
    }
}

# ── i18n ──────────────────────────────────────────────────────────────────────
try:
    from i18n import T, LANG
except ImportError:
    def T(key, lang='en', **kw): return key

# ── Cross-platform opener ─────────────────────────────────────────────────────
import subprocess as _sp
def _open_path(path):
    try:
        if sys.platform == 'win32': os.startfile(path)
        elif sys.platform == 'darwin': _sp.Popen(['open', path])
        else: _sp.Popen(['xdg-open', path])
    except Exception: pass

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, encoding='utf-8') as f: cfg = json.load(f)
            for k,v in DEFAULT_CONFIG.items():
                if k not in cfg: cfg[k] = v
            return cfg
        except: pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uid INTEGER, nama TEXT,
        timestamp TEXT UNIQUE,
        punch INTEGER, pulled_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        uid INTEGER PRIMARY KEY, nama TEXT,
        card_id TEXT, updated_at TEXT
    )''')
    # Pull sessions — history tarikan dari mesin
    c.execute('''CREATE TABLE IF NOT EXISTS pull_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pulled_at TEXT,
        record_count INTEGER,
        new_count INTEGER,
        device_ip TEXT,
        note TEXT
    )''')
    # Excel snapshots — disimpan sebagai blob di DB
    c.execute('''CREATE TABLE IF NOT EXISTS excel_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER,
        created_at TEXT,
        label TEXT,
        filter_year INTEGER,
        filter_month INTEGER,
        data BLOB
    )''')
    conn.commit(); conn.close()

def db_insert_attendance(rows):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ins = 0
    for r in rows:
        ts = r['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if hasattr(r['timestamp'],'strftime') else str(r['timestamp'])
        try:
            c.execute('INSERT INTO attendance (uid,nama,timestamp,punch,pulled_at) VALUES (?,?,?,?,?)',
                      (r['uid'], r['nama'], ts, r['punch'], now))
            ins += 1
        except sqlite3.IntegrityError: pass
    conn.commit(); conn.close()
    return ins

def db_query_attendance(year=None, month=None):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    q = "SELECT uid,nama,timestamp,punch FROM attendance WHERE strftime('%Y',timestamp)>'2000'"
    args = []
    if year:  q += " AND strftime('%Y',timestamp)=?";  args.append(str(year))
    if month: q += " AND strftime('%m',timestamp)=?";  args.append(f"{month:02d}")
    q += " ORDER BY timestamp"
    c.execute(q, args)
    rows = [{'uid':r[0],'nama':r[1],
             'timestamp':datetime.strptime(r[2],'%Y-%m-%d %H:%M:%S'),'punch':r[3]}
            for r in c.fetchall()]
    conn.close(); return rows

def db_count():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM attendance WHERE strftime('%Y',timestamp)>'2000'")
    n = c.fetchone()[0]; conn.close(); return n

def db_upsert_users(users):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for u in users:
        c.execute('INSERT OR REPLACE INTO users (uid,nama,card_id,updated_at) VALUES (?,?,?,?)',
                  (u['uid'], u['nama'], u.get('card_id',''), now))
    conn.commit(); conn.close()

def db_add_pull_session(record_count, new_count, device_ip, note=''):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('INSERT INTO pull_sessions (pulled_at,record_count,new_count,device_ip,note) VALUES (?,?,?,?,?)',
              (now, record_count, new_count, device_ip, note))
    sid = c.lastrowid
    conn.commit(); conn.close()
    return sid

def db_get_pull_sessions():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id,pulled_at,record_count,new_count,device_ip FROM pull_sessions ORDER BY id DESC')
    rows = c.fetchall(); conn.close()
    return rows

def db_delete_pull_session(sid):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM pull_sessions WHERE id=?', (sid,))
    c.execute('DELETE FROM excel_snapshots WHERE session_id=?', (sid,))
    conn.commit(); conn.close()

def db_save_excel_snapshot(session_id, label, year, month, data_bytes):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('INSERT INTO excel_snapshots (session_id,created_at,label,filter_year,filter_month,data) VALUES (?,?,?,?,?,?)',
              (session_id, now, label, year, month or 0, data_bytes))
    sid = c.lastrowid
    conn.commit(); conn.close()
    return sid

def db_get_excel_snapshots():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''SELECT e.id, e.label, e.created_at, e.filter_year, e.filter_month,
                        p.pulled_at, p.record_count
                 FROM excel_snapshots e
                 LEFT JOIN pull_sessions p ON e.session_id=p.id
                 ORDER BY e.id DESC''')
    rows = c.fetchall(); conn.close()
    return rows

def db_load_excel_snapshot(snap_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT data,label FROM excel_snapshots WHERE id=?', (snap_id,))
    row = c.fetchone(); conn.close()
    return row  # (bytes, label)

def db_delete_excel_snapshot(snap_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM excel_snapshots WHERE id=?', (snap_id,))
    conn.commit(); conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR — returns bytes instead of saving to file
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_bytes(rows, cfg):
    """Generate Excel and return as bytes (not saved to disk)."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl tidak terinstall.")

    jam_masuk_std  = cfg.get('jam_masuk','08:00')
    jam_keluar_std = cfg.get('jam_keluar','16:00')
    toleransi      = int(cfg.get('toleransi', 15))
    user_map       = {int(k):v for k,v in cfg.get('user_map',{}).items()}
    lang           = cfg.get('lang','en')

    def F(h): return PatternFill('solid', fgColor=h)
    def Bs(st='thin', co='FFB0B0B0'):
        s=Side(style=st,color=co); return Border(left=s,right=s,top=s,bottom=s)
    def Bm(co='FF888888'):
        s=Side(style='medium',color=co); return Border(left=s,right=s,top=s,bottom=s)
    def fnt(sz=9, bold=False, color='FF000000', italic=False):
        return Font(name='Arial', size=sz, bold=bold, color=color, italic=italic)
    C  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    CW = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L  = Alignment(horizontal='left',   vertical='center')
    R  = Alignment(horizontal='right',  vertical='center')

    CH='FF1A56DB'; CS='FF3B82F6'; CA='FF1E40AF'
    GBG='FFD1FAE5'; GTX='FF065F46'
    RBG='FFFEE2E2'; RTX='FF991B1B'
    YBG='FFFEF9C3'; YTX='FF854D0E'
    OBG='FFFED7AA'; OTX='FF9A3412'
    R1='FFFFFFFF'; R2='FFF0F4FF'; STR='FFE8F0FE'; BD='FFCBD5E1'; HT='FFFFFFFF'

    # month names
    MNAMES = (["","January","February","March","April","May","June",
               "July","August","September","October","November","December"]
              if lang=='en' else
              ["","Januari","Februari","Maret","April","Mei","Juni",
               "Juli","Agustus","September","Oktober","November","Desember"])
    DNAMES = (["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
              if lang=='en' else
              ["Sn","Sl","Rb","Km","Jm","Sb","Mg"])
    DFULL  = (["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
              if lang=='en' else
              ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"])

    # ── build data ────────────────────────────────────────────────────────────
    def _parse_ts(t):
        if isinstance(t, datetime): return t
        if isinstance(t, str):
            for fmt in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d %H:%M'):
                try: return datetime.strptime(t, fmt)
                except: pass
        return None

    _raw = {}
    for r in rows:
        ts = _parse_ts(r['timestamp'])
        if ts is None or ts.year <= 2000: continue
        nama = user_map.get(r['uid'], r.get('nama', f"UID:{r['uid']}"))
        tgl  = ts.date()
        _raw.setdefault((nama, tgl), []).append(ts)

    if not _raw: raise RuntimeError("No data for selected period.")

    class _Row:
        __slots__ = ['nama','tanggal','masuk','keluar','tap','tap_total',
                     'jam_masuk','jam_keluar','terlambat','lembur']

    daily_list = []
    for (nama, tgl), taps in sorted(_raw.items()):
        taps_s = sorted(taps)
        row = _Row()
        row.nama       = nama;   row.tanggal   = tgl
        row.masuk      = taps_s[0]; row.keluar = taps_s[-1]
        row.tap        = 2 if len(taps_s)>1 else 1
        row.tap_total  = len(taps_s)
        row.jam_masuk  = taps_s[0].strftime('%H:%M')
        row.jam_keluar = taps_s[-1].strftime('%H:%M') if len(taps_s)>1 else '-'
        row.terlambat  = 0; row.lembur = 0
        daily_list.append(row)

    class _DF:
        def __init__(self, lst): self._lst = lst
        def __iter__(self): return iter(self._lst)
        def __len__(self): return len(self._lst)
        def min_date(self): return min(r.tanggal for r in self._lst)
        def max_date(self): return max(r.tanggal for r in self._lst)
        def names(self): return sorted(set(r.nama for r in self._lst))
        def months(self): return sorted(set(r.tanggal.strftime('%Y-%m') for r in self._lst))
        def filter_month(self,m): return _DF([r for r in self._lst if r.tanggal.strftime('%Y-%m')==m])
        def filter_name(self,n):  return _DF([r for r in self._lst if r.nama==n])
        def filter_date(self,d):  return _DF([r for r in self._lst if r.tanggal==d])
        def filter_late(self):    return _DF([r for r in self._lst if r.terlambat>0])
        def sort_by(self,*keys):  return _DF(sorted(self._lst, key=lambda r: tuple(getattr(r,k) for k in keys)))

    daily = _DF(daily_list)

    std_in  = datetime.strptime(jam_masuk_std,'%H:%M')
    tol_dt  = std_in + timedelta(minutes=toleransi)
    std_out = datetime.strptime(jam_keluar_std,'%H:%M')

    for row in daily:
        try:
            t = datetime.strptime(row.jam_masuk,'%H:%M')
            row.terlambat = int((t-std_in).total_seconds()//60) if t>tol_dt else 0
        except: row.terlambat = 0
        try:
            if row.jam_keluar=='-': row.lembur=0
            else:
                t = datetime.strptime(row.jam_keluar,'%H:%M')
                row.lembur = int((t-std_out).total_seconds()//60) if t>std_out else 0
        except: row.lembur=0

    months    = daily.months()
    all_names = daily.names()
    wb = Workbook()
    wb.remove(wb.active)

    # ── KARTU ABSENSI per BULAN ───────────────────────────────────────────────
    for month in months:
        yr2,mo2   = int(month[:4]),int(month[5:])
        _,days_in = calendar.monthrange(yr2,mo2)
        bln_label = f"{MNAMES[mo2]} {yr2}"
        mdata     = daily.filter_month(month)

        ws = wb.create_sheet(f'{MNAMES[mo2][:3]}{yr2}')
        ws.sheet_view.showGridLines=False
        ws.page_setup.orientation='landscape'; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1

        TOT_COL=days_in+3; LAT_COL=days_in+4; OT_COL=days_in+5; AVG_COL=days_in+6
        LC=get_column_letter(days_in+6)

        ws.merge_cells(f'A1:{LC}1')
        ws['A1']=('CV REJEKI AMERTA JAYA — ATTENDANCE CARD' if lang=='en'
                  else 'CV REJEKI AMERTA JAYA — KARTU ABSENSI KARYAWAN')
        ws['A1'].font=fnt(13,True,HT); ws['A1'].fill=F(CH); ws['A1'].alignment=C
        ws.row_dimensions[1].height=22

        ws.merge_cells(f'A2:{LC}2')
        _std = ('Check-in' if lang=='en' else 'Masuk Std')
        _tol = ('Tolerance' if lang=='en' else 'Toleransi')
        ws['A2']=f"{bln_label.upper()}  |  {_std}: {jam_masuk_std}  |  {_tol}: {toleransi} min"
        ws['A2'].font=fnt(9,False,HT,italic=True); ws['A2'].fill=F(CS); ws['A2'].alignment=C
        ws.row_dimensions[2].height=15

        ws.merge_cells(f'A3:{LC}3')
        _printed = 'Printed:' if lang=='en' else 'Dicetak:'
        ws['A3']=f"{_printed} {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws['A3'].font=fnt(8,False,'FF555555',italic=True); ws['A3'].fill=F('FFF8FAFF'); ws['A3'].alignment=R
        ws.row_dimensions[3].height=13

        ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=15
        for d in range(1,days_in+1): ws.column_dimensions[get_column_letter(d+2)].width=5
        ws.column_dimensions[get_column_letter(TOT_COL)].width=5.5
        ws.column_dimensions[get_column_letter(LAT_COL)].width=9
        ws.column_dimensions[get_column_letter(OT_COL)].width=8
        ws.column_dimensions[get_column_letter(AVG_COL)].width=7

        HDR=4; ws.row_dimensions[HDR].height=28
        def hc(col,val,bg=CA):
            c=ws.cell(row=HDR,column=col,value=val)
            c.font=fnt(8,True,HT); c.fill=F(bg); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
        hc(1,'No'); hc(2,'Name' if lang=='en' else 'Nama')
        for d in range(1,days_in+1):
            dt=datetime(yr2,mo2,d); dn=DNAMES[dt.weekday()]
            c=ws.cell(row=HDR,column=d+2,value=f'{d}\n{dn}')
            c.font=fnt(7,True,HT); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
            if   dt.weekday()==6: c.fill=F('FF991B1B')
            elif dt.weekday()==5: c.fill=F('FF92400E')
            else:                 c.fill=F(CA)
        _lat_lbl='Late\n(min)' if lang=='en' else 'Terlambat\n(mnt)'
        _ot_lbl ='OT\n(min)'   if lang=='en' else 'Lembur\n(mnt)'
        _avg_lbl='Avg\nIn'     if lang=='en' else 'Rata\nMasuk'
        hc(TOT_COL,'∑\nHadir','FF065F46')
        hc(LAT_COL,_lat_lbl,'FF7C2D12')
        hc(OT_COL, _ot_lbl, 'FF1E40AF')
        hc(AVG_COL,_avg_lbl,'FF1E3A5F')

        for idx,name in enumerate(all_names,1):
            r=HDR+idx; ws.row_dimensions[r].height=17
            sub=mdata.filter_name(name)
            c=ws.cell(row=r,column=1,value=idx)
            c.font=fnt(8); c.fill=F(STR); c.alignment=C; c.border=Bs('thin',BD)
            c=ws.cell(row=r,column=2,value=name)
            c.font=fnt(9,True,'FF1E3A5F'); c.fill=F(STR); c.alignment=L; c.border=Bm('FF93C5FD')

            hadir=0; masuk_list=[]; total_late=0; total_ot=0
            for d in range(1,days_in+1):
                col=d+2; dt=datetime(yr2,mo2,d)
                dr=sub.filter_date(date(yr2,mo2,d))
                if len(dr)>0:
                    jam=dr._lst[0].jam_masuk; late=int(dr._lst[0].terlambat); ot=int(dr._lst[0].lembur)
                    c=ws.cell(row=r,column=col,value=jam)
                    masuk_list.append(jam); hadir+=1; total_late+=late; total_ot+=ot
                    if   late>0:          c.font=fnt(7,True,OTX); c.fill=F(OBG)
                    elif dt.weekday()==6: c.font=fnt(7,True,'FF7C3AED'); c.fill=F('FFEDE9FE')
                    elif dt.weekday()==5: c.font=fnt(7,True,'FF92400E'); c.fill=F(YBG)
                    else:                 c.font=fnt(7,False,GTX); c.fill=F(GBG)
                else:
                    c=ws.cell(row=r,column=col,value='')
                    if   dt.weekday()==6: c.fill=F('FFFCE7F3')
                    elif dt.weekday()==5: c.fill=F(YBG)
                    else:                 c.fill=F(RBG)
                    c.font=fnt(7)
                c.alignment=C; c.border=Bs('thin',BD)

            pct=round(hadir/days_in*100)
            c=ws.cell(row=r,column=TOT_COL,value=hadir)
            c.font=fnt(9,True,GTX if pct>=80 else RTX); c.fill=F(GBG if pct>=80 else RBG)
            c.alignment=C; c.border=Bm()
            c=ws.cell(row=r,column=LAT_COL,value=total_late if total_late else '-')
            c.font=fnt(9,total_late>0,OTX if total_late>0 else 'FF888888')
            c.fill=F(OBG if total_late>0 else R1); c.alignment=C; c.border=Bs('thin',BD)
            c=ws.cell(row=r,column=OT_COL,value=total_ot if total_ot else '-')
            c.font=fnt(9,total_ot>0,'FF1D4ED8' if total_ot>0 else 'FF888888')
            c.fill=F('FFE0EAFF' if total_ot>0 else R1); c.alignment=C; c.border=Bs('thin',BD)
            try:
                if masuk_list:
                    ts2=sum(datetime.strptime(t,'%H:%M').hour*3600+datetime.strptime(t,'%H:%M').minute*60 for t in masuk_list)
                    av=ts2//len(masuk_list); avg=f"{av//3600:02d}:{(av%3600)//60:02d}"
                else: avg='-'
            except: avg='-'
            c=ws.cell(row=r,column=AVG_COL,value=avg)
            c.font=fnt(8,False,'FF1E3A5F'); c.fill=F('FFE0EAFF'); c.alignment=C; c.border=Bs('thin',BD)

        # total harian
        rt=HDR+len(all_names)+1; ws.row_dimensions[rt].height=15
        ws.merge_cells(f'A{rt}:B{rt}')
        _tot_lbl='DAILY TOTAL' if lang=='en' else 'TOTAL HADIR HARIAN'
        c=ws.cell(row=rt,column=1,value=_tot_lbl)
        c.font=fnt(8,True,HT); c.fill=F(CA); c.alignment=C; c.border=Bs('thin',BD)
        for d in range(1,days_in+1):
            col=d+2; cnt=len(set(r.nama for r in mdata.filter_date(date(yr2,mo2,d))))
            c=ws.cell(row=rt,column=col,value=cnt if cnt else '')
            c.font=fnt(8,True,GTX if cnt else 'FFAAAAAA')
            c.fill=F(GBG if cnt else 'FFF1F5F9'); c.alignment=C; c.border=Bs('thin',BD)
        for col in (TOT_COL,LAT_COL,OT_COL,AVG_COL):
            c=ws.cell(row=rt,column=col,value=''); c.fill=F(CA); c.border=Bs('thin',BD)

        # legenda
        rl=rt+2; ws.row_dimensions[rl].height=13
        _legs = ([('On time',GBG,GTX),('Late',OBG,OTX),('Saturday',YBG,'FF92400E'),
                  ('Sunday','FFEDE9FE','FF7C3AED'),('Absent',RBG,RTX)]
                 if lang=='en' else
                 [('Tepat waktu',GBG,GTX),('Terlambat',OBG,OTX),('Sabtu',YBG,'FF92400E'),
                  ('Minggu','FFEDE9FE','FF7C3AED'),('Tidak hadir',RBG,RTX)])
        cl=1
        for lb,bg,tc in _legs:
            ws.merge_cells(start_row=rl,start_column=cl,end_row=rl,end_column=cl+2)
            c=ws.cell(row=rl,column=cl,value=f'■ {lb}')
            c.font=fnt(8,False,tc); c.fill=F(bg); c.alignment=L; cl+=3

    # ── REKAP ─────────────────────────────────────────────────────────────────
    wr=wb.create_sheet('Recap' if lang=='en' else 'Rekap'); wr.sheet_view.showGridLines=False
    for i,w in enumerate([5,15,15,8,8,10,8,12,10,10],1):
        wr.column_dimensions[get_column_letter(i)].width=w
    wr.merge_cells('A1:J1')
    _rh=('ATTENDANCE RECAP — CV REJEKI AMERTA JAYA' if lang=='en'
         else 'REKAP ABSENSI — CV REJEKI AMERTA JAYA')
    wr['A1']=_rh; wr['A1'].font=fnt(13,True,HT); wr['A1'].fill=F(CH); wr['A1'].alignment=C
    wr.row_dimensions[1].height=24
    wr.merge_cells('A2:J2')
    _tmin=daily.min_date().strftime('%d %B %Y'); _tmax=daily.max_date().strftime('%d %B %Y')
    _std_lbl='Std Check-in' if lang=='en' else 'Jam Masuk Std'
    _tol_lbl='Tolerance' if lang=='en' else 'Toleransi'
    wr['A2']=f"Period: {_tmin} - {_tmax}  |  {_std_lbl}: {jam_masuk_std}  |  {_tol_lbl}: {toleransi} min"
    wr['A2'].font=fnt(9,False,'FF444444',italic=True); wr['A2'].alignment=C
    wr['A2'].fill=F('FFF0F4FF'); wr.row_dimensions[2].height=15

    _rhdrs=(['No','Name','Month','Days','Present','Absent','% Present','Late (min)','OT (min)','Status']
            if lang=='en' else
            ['No','Nama','Bulan','Hari','Hadir','Tdk Hadir','% Hadir','Terlambat (mnt)','Lembur (mnt)','Status'])
    for i,h in enumerate(_rhdrs,1):
        c=wr.cell(row=3,column=i,value=h)
        c.font=fnt(9,True,HT); c.fill=F(CA); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
    wr.row_dimensions[3].height=22
    r=4; no=1
    for name in all_names:
        sub=daily.filter_name(name)
        for month in months:
            yr2,mo2=int(month[:4]),int(month[5:])
            _,days_in=calendar.monthrange(yr2,mo2)
            md=sub.filter_month(month)
            if len(md)==0: continue
            hadir=len(md); tidak=days_in-hadir; pct=round(hadir/days_in*100)
            tl=sum(r2.terlambat for r2 in md); to=sum(r2.lembur for r2 in md)
            if lang=='en': status='Good' if pct>=90 else ('Fair' if pct>=75 else 'Poor')
            else:          status='Baik' if pct>=90 else ('Cukup' if pct>=75 else 'Kurang')
            bg=F(R1) if r%2==1 else F(R2)
            vals=[no,name,f"{MNAMES[mo2]} {yr2}",days_in,hadir,tidak,f"{pct}%",
                  tl if tl else '-',to if to else '-',status]
            for i,v in enumerate(vals,1):
                c=wr.cell(row=r,column=i,value=v)
                c.font=fnt(9,i==2); c.fill=bg; c.alignment=C if i!=2 else L; c.border=Bs('thin',BD)
            pc=wr.cell(row=r,column=7); sc=wr.cell(row=r,column=10)
            if pct>=90:
                for x in (pc,sc): x.font=fnt(9,True,GTX); x.fill=F(GBG)
            elif pct>=75:
                for x in (pc,sc): x.font=fnt(9,True,YTX); x.fill=F(YBG)
            else:
                for x in (pc,sc): x.font=fnt(9,True,RTX); x.fill=F(RBG)
            lc=wr.cell(row=r,column=8)
            if tl>0: lc.font=fnt(9,True,OTX); lc.fill=F(OBG)
            wr.row_dimensions[r].height=17; r+=1; no+=1

    # ── LOG DETAIL ────────────────────────────────────────────────────────────
    wd=wb.create_sheet('Log Detail'); wd.sheet_view.showGridLines=False
    for i,w in enumerate([5,14,12,10,10,10,8,10,10],1):
        wd.column_dimensions[get_column_letter(i)].width=w
    wd.merge_cells('A1:I1')
    _lh=('ATTENDANCE LOG DETAIL — CV REJEKI AMERTA JAYA' if lang=='en'
         else 'LOG DETAIL ABSENSI — CV REJEKI AMERTA JAYA')
    wd['A1']=_lh; wd['A1'].font=fnt(12,True,HT); wd['A1'].fill=F(CH); wd['A1'].alignment=C
    wd.row_dimensions[1].height=22
    _lhdrs=(['No','Name','Date','Day','Check-in','Check-out','Taps','Late','OT']
            if lang=='en' else
            ['No','Nama','Tanggal','Hari','Jam Masuk','Jam Keluar','Tap','Terlambat','Lembur'])
    for i,h in enumerate(_lhdrs,1):
        c=wd.cell(row=2,column=i,value=h)
        c.font=fnt(9,True,HT); c.fill=F(CA); c.alignment=C; c.border=Bs('thin','FF93C5FD')
    wd.row_dimensions[2].height=18
    ds=daily.sort_by('tanggal','nama')
    for i,row in enumerate(ds):
        r3=i+3; tgl=row.tanggal
        dn=DFULL[tgl.weekday()]; bg=F(R1) if i%2==0 else F(R2)
        late=int(row.terlambat); ot=int(row.lembur)
        tap_info=f"{row.tap_total}x" if hasattr(row,'tap_total') else '-'
        vals=[i+1,row.nama,tgl.strftime('%d/%m/%Y'),dn,
              row.jam_masuk,row.jam_keluar,tap_info,
              f"{late} min" if late>0 else '-',
              f"{ot} min"   if ot>0   else '-']
        for col,v in enumerate(vals,1):
            c=wd.cell(row=r3,column=col,value=v)
            c.font=fnt(9,col==2)
            if col==8 and late>0: c.font=fnt(9,True,OTX); c.fill=F(OBG)
            elif col==9 and ot>0: c.font=fnt(9,True,'FF1D4ED8'); c.fill=F('FFE0EAFF')
            elif col==7 and hasattr(row,'tap_total') and row.tap_total>2:
                c.font=fnt(9,True,'FF6B21A8'); c.fill=F('FFEDE9FE')
            else: c.fill=bg
            c.alignment=C if col!=2 else L; c.border=Bs('thin',BD)
        wd.row_dimensions[r3].height=15

    ms=[s for s in wb.sheetnames if s not in ('Recap','Rekap','Log Detail')]
    wb._sheets=[wb[s] for s in ms+[s for s in ('Recap','Rekap','Log Detail') if s in wb.sheetnames]]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel_for_preview(data_bytes, sheet_idx=0):
    """Parse Excel bytes and return (headers, rows) for Treeview display."""
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[sheet_idx]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows: return [], []
        # skip merged title rows (usually rows 1-3), find header row
        # header = first row where most cells are non-None
        hdr_idx = 0
        for i, row in enumerate(all_rows):
            non_none = sum(1 for c in row if c is not None)
            if non_none >= 3:
                hdr_idx = i
                break
        headers = [str(c) if c is not None else '' for c in all_rows[hdr_idx]]
        # remove empty trailing cols
        while headers and not headers[-1]: headers.pop()
        n_cols = len(headers)
        data_rows = []
        for row in all_rows[hdr_idx+1:]:
            r = [str(c) if c is not None else '' for c in row[:n_cols]]
            if any(c for c in r): data_rows.append(r)
        wb.close()
        return headers, data_rows
    except Exception as e:
        return ['Error'], [[str(e)]]


# ─────────────────────────────────────────────────────────────────────────────
# DIALOGS
# ─────────────────────────────────────────────────────────────────────────────
class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, cfg, on_save):
        super().__init__(parent)
        self.title("Settings"); self.resizable(False,False); self.grab_set()
        self.cfg=cfg; self.on_save=on_save; self._build()

    def _build(self):
        nb=ttk.Notebook(self); nb.pack(fill='both',expand=True,padx=10,pady=10)
        t1=ttk.Frame(nb); nb.add(t1,text='Connection & Hours')
        fields=[('IP Address','ip'),('Port','port'),
                ('Standard Check-in (HH:MM)','jam_masuk'),
                ('Standard Check-out (HH:MM)','jam_keluar'),
                ('Late Tolerance (minutes)','toleransi')]
        self.vars={}
        for i,(label,key) in enumerate(fields):
            ttk.Label(t1,text=label).grid(row=i,column=0,sticky='w',padx=10,pady=4)
            v=tk.StringVar(value=str(self.cfg.get(key,'')))
            ttk.Entry(t1,textvariable=v,width=20).grid(row=i,column=1,sticky='w',padx=10,pady=4)
            self.vars[key]=v
        ab=tk.BooleanVar(value=self.cfg.get('auto_backup',False))
        ttk.Checkbutton(t1,text='Auto backup CSV after pull',variable=ab).grid(
            row=len(fields),column=0,columnspan=2,sticky='w',padx=10,pady=4)
        self.vars['auto_backup']=ab

        t2=ttk.Frame(nb); nb.add(t2,text='Staff Names')
        ttk.Label(t2,text='Format: UID=Name (one per line)',foreground='#666').pack(anchor='w',padx=10,pady=(8,2))
        self.user_text=scrolledtext.ScrolledText(t2,width=35,height=14,font=('Consolas',9))
        self.user_text.pack(padx=10,pady=(0,8))
        um=self.cfg.get('user_map',{})
        self.user_text.insert('end','\n'.join(f"{k}={v}" for k,v in sorted(um.items(),key=lambda x:int(x[0]))))

        bf=tk.Frame(self); bf.pack(pady=(0,10))
        ttk.Button(bf,text='💾 Save',command=self._save).pack(side='left',padx=6)
        ttk.Button(bf,text='Cancel',command=self.destroy).pack(side='left',padx=6)

    def _save(self):
        for key,v in self.vars.items():
            try:
                if key=='toleransi': self.cfg[key]=int(v.get())
                elif key=='auto_backup': self.cfg[key]=bool(v.get())
                else: self.cfg[key]=v.get()
            except: self.cfg[key]=v.get()
        um={}
        for line in self.user_text.get('1.0','end').strip().splitlines():
            if '=' in line:
                k,_,v2=line.partition('=')
                if k.strip().isdigit(): um[k.strip()]=v2.strip()
        self.cfg['user_map']=um
        save_config(self.cfg)
        self.on_save(self.cfg)
        self.destroy()


class DeviceInfoDialog(tk.Toplevel):
    def __init__(self, parent, info_dict):
        super().__init__(parent)
        self.title("Device Info"); self.resizable(False,False); self.grab_set()
        tk.Label(self,text="  Device Info — eFace10",font=("Segoe UI",11,"bold"),
                 bg='#1A56DB',fg='white').pack(fill='x',pady=4)
        fr=ttk.Frame(self,padding=12); fr.pack(fill='both',expand=True)
        for i,(k,v) in enumerate(info_dict.items()):
            ttk.Label(fr,text=k+':',font=('Segoe UI',9,'bold')).grid(row=i,column=0,sticky='w',pady=3,padx=(0,16))
            ttk.Label(fr,text=str(v),font=('Segoe UI',9)).grid(row=i,column=1,sticky='w')
        ttk.Button(self,text='Close',command=self.destroy).pack(pady=8)


class UserManagerDialog(tk.Toplevel):
    def __init__(self, parent, users):
        super().__init__(parent)
        self.title("Users on Device"); self.resizable(False,False); self.grab_set()
        tk.Label(self,text="  Users registered on eFace10",font=("Segoe UI",11,"bold"),
                 bg='#1A56DB',fg='white').pack(fill='x',pady=4)
        fr=ttk.Frame(self,padding=8); fr.pack(fill='both',expand=True)
        cols=('UID','Name','Card ID')
        self.tree=ttk.Treeview(fr,columns=cols,show='headings',height=14)
        for col,w in zip(cols,[80,180,140]):
            self.tree.heading(col,text=col); self.tree.column(col,width=w)
        sb=ttk.Scrollbar(fr,orient='vertical',command=self.tree.yview)
        self.tree.config(yscrollcommand=sb.set)
        self.tree.pack(side='left'); sb.pack(side='right',fill='y')
        for u in users:
            self.tree.insert('','end',values=(u['uid'],u['nama'],u.get('card_id','')))
        bf=tk.Frame(self); bf.pack(pady=6)
        ttk.Label(bf,text=f"Total: {len(users)} users").pack(side='left',padx=10)
        ttk.Button(bf,text='Close',command=self.destroy).pack(side='right',padx=10)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP — SPLIT PANEL
# ─────────────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"ZKTeco eFace10 Utility v{APP_VERSION} — CV RAJ")
        self.resizable(True,True)
        self.minsize(1100,680)
        self.configure(bg='#F1F5F9')
        self.cfg=load_config()
        self._cache=[]
        self._current_snap_bytes=None   # Excel bytes di memory
        self._current_snap_id=None
        init_db()
        self._build_ui()
        self._update_status()
        self._refresh_history()

    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────────────────
        hdr=tk.Frame(self,bg='#1A56DB'); hdr.pack(fill='x')
        _base=os.path.dirname(os.path.abspath(__file__))
        try:
            if sys.platform=='win32' and os.path.exists(os.path.join(_base,'app_icon.ico')):
                self.iconbitmap(os.path.join(_base,'app_icon.ico'))
            elif os.path.exists(os.path.join(_base,'app_icon.png')):
                _img=tk.PhotoImage(file=os.path.join(_base,'app_icon.png'))
                self.iconphoto(True,_img)
        except Exception: pass

        tk.Label(hdr,text=f"  ZKTeco eFace10 Utility  ·  CV RAJ",
                 font=('Segoe UI',12,'bold'),bg='#1A56DB',fg='white',pady=8).pack(side='left')
        ttk.Button(hdr,text='⚙ Settings',command=self._open_settings).pack(side='right',padx=8,pady=5)
        ttk.Button(hdr,text='⬆ Update',command=self._check_update).pack(side='right',padx=4,pady=5)
        tk.Label(hdr,text=f"v{APP_VERSION}  ",font=('Segoe UI',8),bg='#1A56DB',fg='#93C5FD').pack(side='right')
        lf=tk.Frame(hdr,bg='#1A56DB'); lf.pack(side='right',padx=4)
        tk.Label(lf,text='🌐',bg='#1A56DB',fg='white',font=('Segoe UI',9)).pack(side='left')
        self.lang_var=tk.StringVar(value=self.cfg.get('lang','en'))
        lc=ttk.Combobox(lf,textvariable=self.lang_var,values=['en','id'],width=4,state='readonly')
        lc.pack(side='left',pady=6,padx=(0,4))
        lc.bind('<<ComboboxSelected>>',self._on_lang_change)

        # ── Main split pane ───────────────────────────────────────────────────
        paned=ttk.PanedWindow(self,orient='horizontal')
        paned.pack(fill='both',expand=True,padx=4,pady=4)

        # ══ LEFT PANEL — Workflow ══════════════════════════════════════════════
        left=tk.Frame(paned,bg='#F1F5F9',width=420)
        left.pack_propagate(False)
        paned.add(left,weight=0)

        pad=dict(padx=8,pady=3)

        # Koneksi
        fc=ttk.LabelFrame(left,text='Device Connection',padding=6)
        fc.pack(fill='x',**pad)
        ttk.Label(fc,text='IP:').grid(row=0,column=0,sticky='w')
        self.ip_var=tk.StringVar(value=self.cfg['ip'])
        ttk.Entry(fc,textvariable=self.ip_var,width=16).grid(row=0,column=1,padx=3)
        ttk.Label(fc,text='Port:').grid(row=0,column=2,sticky='w',padx=(6,0))
        self.port_var=tk.StringVar(value=self.cfg['port'])
        ttk.Entry(fc,textvariable=self.port_var,width=7).grid(row=0,column=3,padx=3)
        ttk.Button(fc,text='🔌 Test',command=lambda:self._run(self._do_test)).grid(row=0,column=4,padx=4)
        ttk.Button(fc,text='ℹ Info',command=lambda:self._run(self._do_info)).grid(row=0,column=5,padx=2)
        ttk.Button(fc,text='👤 Users',command=lambda:self._run(self._do_users)).grid(row=0,column=6,padx=2)
        self.conn_lbl=tk.Label(fc,text='● Not connected',font=('Segoe UI',7),fg='#888',bg='#F1F5F9')
        self.conn_lbl.grid(row=1,column=0,columnspan=7,sticky='w',pady=(2,0))

        # Alur kerja — step boxes
        fw=ttk.LabelFrame(left,text='Workflow',padding=8)
        fw.pack(fill='x',**pad)

        def mkstep(col,no,title,desc):
            f=tk.Frame(fw,bg='#EFF6FF',relief='groove',bd=1)
            f.grid(row=0,column=col,padx=3,pady=2,sticky='n')
            tk.Label(f,text=f"{'①②③④'[no-1]} {title}",font=('Segoe UI',8,'bold'),
                     bg='#EFF6FF',fg='#1e40af').pack(pady=(5,1),padx=8)
            tk.Label(f,text=desc,font=('Segoe UI',7),bg='#EFF6FF',fg='#555').pack(padx=8)
            return f

        def arrow(col):
            tk.Label(fw,text='→',font=('Segoe UI',12),fg='#94A3B8',bg='#F1F5F9').grid(row=0,column=col,padx=1)

        f1=mkstep(0,1,'Set Time','Sync RTC\nto PC')
        self.btn_time=ttk.Button(f1,text='🕐 Set Time',width=13,command=lambda:self._run(self._do_settime))
        self.btn_time.pack(pady=(3,6),padx=6)
        arrow(1)

        f2=mkstep(2,2,'Pull Data','Fetch log\n& save DB')
        self.btn_pull=ttk.Button(f2,text='📥 Pull Data',width=13,command=lambda:self._run(self._do_pull))
        self.btn_pull.pack(pady=(3,6),padx=6)
        arrow(3)

        f3=mkstep(4,3,'Filter','Month &\nyear')
        ff=tk.Frame(f3,bg='#EFF6FF'); ff.pack(padx=4,pady=2)
        now=datetime.now()
        tk.Label(ff,text='Month:',font=('Segoe UI',7),bg='#EFF6FF').grid(row=0,column=0,sticky='w')
        self.bulan_var=tk.StringVar(value='All')
        _months=['All']+['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        ttk.Combobox(ff,textvariable=self.bulan_var,values=_months,width=6,state='readonly').grid(row=0,column=1,padx=2)
        tk.Label(ff,text='Year:',font=('Segoe UI',7),bg='#EFF6FF').grid(row=1,column=0,sticky='w',pady=2)
        self.tahun_var=tk.StringVar(value=str(now.year))
        ttk.Combobox(ff,textvariable=self.tahun_var,values=[str(y) for y in range(now.year-3,now.year+2)],
                     width=6,state='readonly').grid(row=1,column=1,padx=2)
        tk.Label(f3,text='Source:',font=('Segoe UI',7),bg='#EFF6FF').pack(pady=(2,0))
        self.src_var=tk.StringVar(value='database')
        ttk.Radiobutton(f3,text='Database',variable=self.src_var,value='database').pack(anchor='w',padx=6)
        ttk.Radiobutton(f3,text='Cache',variable=self.src_var,value='cache').pack(anchor='w',padx=6)
        arrow(5)

        f4=mkstep(6,4,'Preview','Generate &\nview report')
        self.btn_report=ttk.Button(f4,text='📊 Preview',width=13,command=lambda:self._run(self._do_report))
        self.btn_report.pack(pady=(3,6),padx=6)

        # shortcut bar
        sf=tk.Frame(fw,bg='#F1F5F9'); sf.grid(row=1,column=0,columnspan=7,pady=(5,0))
        self.btn_all=ttk.Button(sf,text='⚡ All at Once',command=lambda:self._run(self._do_all))
        self.btn_all.pack(side='left',padx=3)
        ttk.Separator(sf,orient='vertical').pack(side='left',fill='y',padx=6)
        ttk.Button(sf,text='🗑 Clear Device Log',command=self._confirm_clear).pack(side='left',padx=3)

        # status
        self.data_lbl=tk.Label(fw,text='...',font=('Segoe UI',7),fg='#1e40af',bg='#F1F5F9')
        self.data_lbl.grid(row=2,column=0,columnspan=7,pady=(3,0))

        # Log
        fl=ttk.LabelFrame(left,text='Log',padding=4)
        fl.pack(fill='both',expand=True,**pad)
        self.log_box=scrolledtext.ScrolledText(fl,font=('Consolas',8),state='disabled',
                                                bg='#0F172A',fg='#CBD5E1',height=8)
        self.log_box.pack(fill='both',expand=True)

        # Status bar (left bottom)
        self.status_var=tk.StringVar(value='Ready.')
        ttk.Label(left,textvariable=self.status_var,relief='sunken',anchor='w',
                  font=('Segoe UI',8)).pack(fill='x',padx=8,pady=(0,4))

        self._btns=[self.btn_time,self.btn_pull,self.btn_report,self.btn_all]

        # ══ RIGHT PANEL — Preview + History ════════════════════════════════════
        right=tk.Frame(paned,bg='#F8FAFC')
        paned.add(right,weight=1)

        nb=ttk.Notebook(right)
        nb.pack(fill='both',expand=True,padx=4,pady=4)

        # ── Tab 1: Report Viewer ──────────────────────────────────────────────
        tab_view=ttk.Frame(nb)
        nb.add(tab_view,text='📊 Report Viewer')

        # toolbar
        vtb=tk.Frame(tab_view,bg='#E2E8F0',pady=4)
        vtb.pack(fill='x')
        tk.Label(vtb,text='Sheet:',font=('Segoe UI',9),bg='#E2E8F0').pack(side='left',padx=(8,2))
        self.sheet_var=tk.StringVar()
        self.sheet_cb=ttk.Combobox(vtb,textvariable=self.sheet_var,width=18,state='readonly')
        self.sheet_cb.pack(side='left',padx=4)
        self.sheet_cb.bind('<<ComboboxSelected>>',self._on_sheet_change)
        ttk.Button(vtb,text='💾 Save to File',command=self._save_excel).pack(side='left',padx=8)
        ttk.Button(vtb,text='🔄 Refresh',command=lambda:self._run(self._do_report)).pack(side='left',padx=2)
        self.snap_lbl=tk.Label(vtb,text='No report loaded',font=('Segoe UI',8),
                                fg='#64748B',bg='#E2E8F0')
        self.snap_lbl.pack(side='right',padx=10)

        # Treeview with scrollbars
        tv_frame=tk.Frame(tab_view)
        tv_frame.pack(fill='both',expand=True)
        self.tree_vsb=ttk.Scrollbar(tv_frame,orient='vertical')
        self.tree_hsb=ttk.Scrollbar(tv_frame,orient='horizontal')
        self.tree_vsb.pack(side='right',fill='y')
        self.tree_hsb.pack(side='bottom',fill='x')
        self.report_tree=ttk.Treeview(tv_frame,show='headings',
                                       yscrollcommand=self.tree_vsb.set,
                                       xscrollcommand=self.tree_hsb.set)
        self.report_tree.pack(fill='both',expand=True)
        self.tree_vsb.config(command=self.report_tree.yview)
        self.tree_hsb.config(command=self.report_tree.xview)

        # alternating row colors
        self.report_tree.tag_configure('odd',  background='#FFFFFF')
        self.report_tree.tag_configure('even', background='#F0F4FF')

        # ── Tab 2: History ────────────────────────────────────────────────────
        tab_hist=ttk.Frame(nb)
        nb.add(tab_hist,text='📋 Pull History')

        # History toolbar
        htb=tk.Frame(tab_hist,bg='#E2E8F0',pady=4)
        htb.pack(fill='x')
        tk.Label(htb,text='Pull sessions from device',font=('Segoe UI',9,'bold'),
                 bg='#E2E8F0').pack(side='left',padx=8)
        ttk.Button(htb,text='🔄 Refresh',command=self._refresh_history).pack(side='right',padx=8)

        # Split history: top=sessions, bottom=snapshots
        hist_paned=ttk.PanedWindow(tab_hist,orient='vertical')
        hist_paned.pack(fill='both',expand=True)

        # Sessions table
        sess_frame=ttk.LabelFrame(hist_paned,text='Pull Sessions',padding=4)
        hist_paned.add(sess_frame,weight=1)
        sess_vsb=ttk.Scrollbar(sess_frame,orient='vertical')
        sess_vsb.pack(side='right',fill='y')
        self.sess_tree=ttk.Treeview(sess_frame,
                                     columns=('id','date','records','new','ip'),
                                     show='headings',height=6,
                                     yscrollcommand=sess_vsb.set)
        sess_vsb.config(command=self.sess_tree.yview)
        for col,w,lbl in [('id',40,'ID'),('date',160,'Pull Date'),
                           ('records',80,'Records'),('new',80,'New'),('ip',120,'Device IP')]:
            self.sess_tree.heading(col,text=lbl)
            self.sess_tree.column(col,width=w,anchor='center')
        self.sess_tree.pack(fill='both',expand=True)
        self.sess_tree.tag_configure('odd',background='#FFFFFF')
        self.sess_tree.tag_configure('even',background='#F0F4FF')

        sbf=tk.Frame(sess_frame); sbf.pack(fill='x',pady=(4,0))
        ttk.Button(sbf,text='📊 Load to Preview',command=self._load_session_to_preview).pack(side='left',padx=4)
        ttk.Button(sbf,text='🗑 Delete Session',command=self._delete_session).pack(side='left',padx=4)
        tk.Label(sbf,text='← Select a session first',font=('Segoe UI',8),fg='#888').pack(side='left',padx=6)

        # Snapshots table
        snap_frame=ttk.LabelFrame(hist_paned,text='Saved Reports (in database)',padding=4)
        hist_paned.add(snap_frame,weight=1)
        snap_vsb=ttk.Scrollbar(snap_frame,orient='vertical')
        snap_vsb.pack(side='right',fill='y')
        self.snap_tree=ttk.Treeview(snap_frame,
                                     columns=('id','label','date','period'),
                                     show='headings',height=6,
                                     yscrollcommand=snap_vsb.set)
        snap_vsb.config(command=self.snap_tree.yview)
        for col,w,lbl in [('id',40,'ID'),('label',200,'Report Label'),
                           ('date',160,'Created'),('period',100,'Period')]:
            self.snap_tree.heading(col,text=lbl)
            self.snap_tree.column(col,width=w,anchor='center' if col!='label' else 'w')
        self.snap_tree.pack(fill='both',expand=True)
        self.snap_tree.tag_configure('odd',background='#FFFFFF')
        self.snap_tree.tag_configure('even',background='#F0F4FF')
        self.snap_tree.bind('<Double-1>',self._load_snapshot)

        snbf=tk.Frame(snap_frame); snbf.pack(fill='x',pady=(4,0))
        ttk.Button(snbf,text='📂 Load Report',command=self._load_snapshot).pack(side='left',padx=4)
        ttk.Button(snbf,text='💾 Export to File',command=self._export_snapshot).pack(side='left',padx=4)
        ttk.Button(snbf,text='🗑 Delete',command=self._delete_snapshot).pack(side='left',padx=4)
        tk.Label(snbf,text='Double-click to preview',font=('Segoe UI',8),fg='#888').pack(side='left',padx=6)

    # ── UI Helpers ────────────────────────────────────────────────────────────
    def _log(self,msg):
        ts=datetime.now().strftime('%H:%M:%S')
        self.log_box.config(state='normal')
        self.log_box.insert('end',f'[{ts}] {msg}\n')
        self.log_box.see('end')
        self.log_box.config(state='disabled')
        self.after(0,lambda:self.status_var.set(msg))

    def _get_conn(self):
        from zk import ZK
        return ZK(self.ip_var.get().strip(),port=int(self.port_var.get()),
                  timeout=15,password=0,force_udp=False,ommit_ping=False).connect()

    def _set_buttons(self,state):
        for b in self._btns:
            try: b.config(state=state)
            except: pass

    def _run(self,fn):
        self._set_buttons('disabled')
        threading.Thread(target=self._worker,args=(fn,),daemon=True).start()

    def _worker(self,fn):
        try: fn()
        except Exception as e:
            self.after(0,lambda:self._log(f'[ERROR] {e}'))
            self.after(0,lambda:messagebox.showerror('Error',str(e)))
        finally:
            self.after(0,lambda:self._set_buttons('normal'))
            self.after(0,self._update_status)

    def _update_status(self):
        n=db_count()
        self.data_lbl.config(text=f'DB: {n:,} records  |  Cache: {len(self._cache):,}')

    def _open_settings(self):
        def on_save(new_cfg):
            self.cfg=new_cfg
            self.ip_var.set(new_cfg['ip'])
            self.port_var.set(new_cfg['port'])
            self._log('✓ Settings saved')
        SettingsDialog(self,self.cfg,on_save)

    def _on_lang_change(self,_=None):
        lang=self.lang_var.get()
        self.cfg['lang']=lang; save_config(self.cfg)
        msg=("Language set to English. Restart to apply." if lang=='en'
             else "Bahasa diubah ke Indonesia. Restart untuk menerapkan.")
        messagebox.showinfo("Language",msg)

    def _confirm_clear(self):
        if messagebox.askyesno('Confirm',
            'Delete ALL attendance log from device memory?\n\n'
            'Data already in local database will NOT be deleted.'):
            self._run(self._do_clear)

    # ── Preview / Viewer ──────────────────────────────────────────────────────
    def _load_excel_to_viewer(self, data_bytes, label=''):
        """Parse Excel bytes and populate Treeview."""
        try:
            from openpyxl import load_workbook
            import io
            wb=load_workbook(io.BytesIO(data_bytes),read_only=True,data_only=True)
            sheet_names=wb.sheetnames
            wb.close()
            self._current_snap_bytes=data_bytes
            self.sheet_cb['values']=sheet_names
            self.sheet_var.set(sheet_names[0] if sheet_names else '')
            self.snap_lbl.config(text=label or 'Report loaded')
            self._render_sheet(data_bytes,0)
        except Exception as e:
            self._log(f'[ERROR] Cannot load preview: {e}')

    def _render_sheet(self, data_bytes, sheet_idx):
        headers,rows=parse_excel_for_preview(data_bytes,sheet_idx)
        # clear
        self.report_tree.delete(*self.report_tree.get_children())
        if not headers: return
        self.report_tree['columns']=headers
        for col in headers:
            self.report_tree.heading(col,text=col)
            w=max(60, min(160, len(col)*11))
            self.report_tree.column(col,width=w,minwidth=40,anchor='center')
        for i,row in enumerate(rows):
            tag='odd' if i%2==0 else 'even'
            self.report_tree.insert('','end',values=row,tags=(tag,))

    def _on_sheet_change(self,_=None):
        if not self._current_snap_bytes: return
        idx=self.sheet_cb['values'].index(self.sheet_var.get())
        threading.Thread(target=lambda:self._render_sheet(self._current_snap_bytes,idx),daemon=True).start()

    def _save_excel(self):
        if not self._current_snap_bytes:
            messagebox.showwarning('No Report','Generate a report first.')
            return
        path=filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel','*.xlsx')],
            initialfile=f'Absensi_CVRAJ_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        if path:
            with open(path,'wb') as f: f.write(self._current_snap_bytes)
            self._log(f'✓ Saved to {path}')
            _open_path(os.path.dirname(path))

    # ── History ───────────────────────────────────────────────────────────────
    def _refresh_history(self):
        # sessions
        self.sess_tree.delete(*self.sess_tree.get_children())
        for i,row in enumerate(db_get_pull_sessions()):
            sid,pulled_at,rec,new_,ip=row
            tag='odd' if i%2==0 else 'even'
            self.sess_tree.insert('','end',iid=str(sid),
                                   values=(sid,pulled_at,f"{rec:,}",f"+{new_:,}",ip),tags=(tag,))
        # snapshots
        self.snap_tree.delete(*self.snap_tree.get_children())
        for i,row in enumerate(db_get_excel_snapshots()):
            snap_id,label,created_at,yr,mo,_,_=row
            period=f"{yr}/{mo:02d}" if mo else str(yr)
            tag='odd' if i%2==0 else 'even'
            self.snap_tree.insert('','end',iid=str(snap_id),
                                   values=(snap_id,label,created_at,period),tags=(tag,))

    def _load_session_to_preview(self):
        sel=self.sess_tree.selection()
        if not sel: messagebox.showwarning('Select','Select a pull session first.'); return
        sid=int(sel[0])
        # query attendance for this session (by pulled_at)
        conn=sqlite3.connect(DB_FILE)
        c=conn.cursor()
        c.execute('SELECT pulled_at FROM pull_sessions WHERE id=?',(sid,))
        row=c.fetchone(); conn.close()
        if not row: return
        pulled_at=row[0]
        # get all records pulled in that session
        conn=sqlite3.connect(DB_FILE)
        c=conn.cursor()
        c.execute("SELECT uid,nama,timestamp,punch FROM attendance WHERE pulled_at=? ORDER BY timestamp",(pulled_at,))
        att_rows=[{'uid':r[0],'nama':r[1],
                   'timestamp':datetime.strptime(r[2],'%Y-%m-%d %H:%M:%S'),'punch':r[3]}
                  for r in c.fetchall()]
        conn.close()
        if not att_rows:
            messagebox.showinfo('Empty','No attendance records found for this session.'); return
        self._run(lambda: self._generate_and_show(att_rows, f"Session #{sid} ({pulled_at[:10]})", sid))

    def _generate_and_show(self, rows, label, session_id=None):
        self._log(f'Generating report for {label} ...')
        yr  =int(self.tahun_var.get())
        bln =self.bulan_var.get()
        mo  =(['All','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'].index(bln)
               if bln!='All' else None)
        if mo: rows=[r for r in rows if r['timestamp'].month==mo and r['timestamp'].year==yr]
        data=generate_excel_bytes(rows,self.cfg)
        # save snapshot to DB
        period=f"{bln} {yr}" if bln!='All' else f"All {yr}"
        snap_id=db_save_excel_snapshot(session_id or 0,f"{label} — {period}",yr,mo,data)
        self._current_snap_id=snap_id
        self.after(0,lambda:self._load_excel_to_viewer(data,f"{label} — {period}"))
        self.after(0,self._refresh_history)
        self._log(f'✓ Report ready — {len(rows)} records | Saved to DB (snapshot #{snap_id})')

    def _load_snapshot(self,_=None):
        sel=self.snap_tree.selection()
        if not sel: messagebox.showwarning('Select','Select a report first.'); return
        snap_id=int(sel[0])
        row=db_load_excel_snapshot(snap_id)
        if not row: return
        data,label=row
        self._current_snap_id=snap_id
        self._load_excel_to_viewer(data,label)
        self._log(f'✓ Loaded snapshot #{snap_id}: {label}')

    def _export_snapshot(self):
        sel=self.snap_tree.selection()
        if not sel:
            if self._current_snap_bytes: self._save_excel(); return
            messagebox.showwarning('Select','Select a report to export.'); return
        snap_id=int(sel[0])
        row=db_load_excel_snapshot(snap_id)
        if not row: return
        data,label=row
        safe=label.replace(' ','_').replace('/','_').replace(':','')[:40]
        path=filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel','*.xlsx')],
            initialfile=f'{safe}.xlsx'
        )
        if path:
            with open(path,'wb') as f: f.write(data)
            self._log(f'✓ Exported to {path}')
            _open_path(os.path.dirname(path))

    def _delete_session(self):
        sel=self.sess_tree.selection()
        if not sel: return
        sid=int(sel[0])
        if messagebox.askyesno('Delete',f'Delete pull session #{sid}?\nAll saved reports for this session will also be deleted.'):
            db_delete_pull_session(sid)
            self._refresh_history()
            self._log(f'✓ Session #{sid} deleted')

    def _delete_snapshot(self):
        sel=self.snap_tree.selection()
        if not sel: return
        snap_id=int(sel[0])
        if messagebox.askyesno('Delete',f'Delete report snapshot #{snap_id}?'):
            db_delete_excel_snapshot(snap_id)
            self._refresh_history()
            self._log(f'✓ Snapshot #{snap_id} deleted')

    # ── Device actions ────────────────────────────────────────────────────────
    def _do_test(self):
        ip=self.ip_var.get().strip()
        self._log(f'Testing connection to {ip}:{self.port_var.get()} ...')
        conn=self._get_conn()
        try:
            fw=conn.get_firmware_version()
            self._log(f'✓ Connected! Firmware: {fw}')
            self.after(0,lambda:self.conn_lbl.config(text=f'● {ip}',fg='#16a34a'))
        finally: conn.disconnect()

    def _do_info(self):
        self._log('Fetching device info ...')
        conn=self._get_conn()
        try:
            info={'Serial Number':conn.get_serialnumber(),'Firmware':conn.get_firmware_version(),
                  'Device Time':str(conn.get_time()),
                  'Users':len(conn.get_users()),'Logs':len(conn.get_attendance()),
                  'Local DB':f'{db_count():,} records'}
            self._log('✓ Device info received')
            self.after(0,lambda:DeviceInfoDialog(self,info))
        finally: conn.disconnect()

    def _do_users(self):
        self._log('Fetching users from device ...')
        conn=self._get_conn()
        try:
            users=conn.get_users()
            ulist=[{'uid':int(u.user_id),'nama':u.name,'card_id':getattr(u,'card','') or ''} for u in users]
            db_upsert_users(ulist)
            for u in ulist: self.cfg['user_map'][str(u['uid'])]=u['nama']
            save_config(self.cfg)
            self._log(f'✓ {len(ulist)} users found and synced to config')
            self.after(0,lambda:UserManagerDialog(self,ulist))
        finally: conn.disconnect()

    def _do_settime(self):
        self._log('Setting device time ...')
        conn=self._get_conn()
        try:
            now=datetime.now(); conn.set_time(now)
            self._log(f'✓ Device time → {now.strftime("%Y-%m-%d %H:%M:%S")}')
        finally: conn.disconnect()

    def _do_pull(self):
        self._log('Pulling attendance data from device ...')
        conn=self._get_conn()
        try:
            atts=conn.get_attendance()
            if not atts: self._log('⚠ No data on device.'); return
            um={int(k):v for k,v in self.cfg.get('user_map',{}).items()}
            self._cache=[{'uid':int(a.user_id),'nama':um.get(int(a.user_id),f'UID:{a.user_id}'),
                          'timestamp':a.timestamp,'punch':a.punch} for a in atts]
            new=db_insert_attendance(self._cache)
            sid=db_add_pull_session(len(atts),new,self.ip_var.get().strip())
            self._log(f'✓ {len(atts)} records pulled  |  {new} new saved to database')
            self._log(f'  Pull session #{sid} recorded')
            if self.cfg.get('auto_backup',False):
                ts=datetime.now().strftime('%Y%m%d_%H%M%S')
                path=os.path.join(os.path.dirname(os.path.abspath(__file__)),f'backup_raw_{ts}.csv')
                with open(path,'w',newline='',encoding='utf-8') as f:
                    w=csv.writer(f); w.writerow(['UserID','Name','Timestamp','Status'])
                    for a in atts:
                        uid=int(a.user_id)
                        w.writerow([uid,um.get(uid,f'UID:{uid}'),
                                    a.timestamp.strftime('%Y-%m-%d %H:%M:%S') if a.timestamp else '',
                                    'Check-In' if a.punch==0 else 'Check-Out'])
                self._log(f'  Backup CSV → {path}')
            self.after(0,self._refresh_history)
        finally: conn.disconnect()

    def _do_clear(self):
        self._log('Clearing log from device memory ...')
        conn=self._get_conn()
        try:
            conn.clear_attendance()
            self._log('✓ Device attendance log cleared')
        finally: conn.disconnect()

    def _do_report(self):
        yr  =int(self.tahun_var.get())
        bln =self.bulan_var.get()
        src =self.src_var.get()
        _month_list=['All','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        mo  =_month_list.index(bln) if bln!='All' else None
        if src=='cache':
            if not self._cache: raise RuntimeError('No cache. Pull data first.')
            rows=self._cache
        else:
            rows=db_query_attendance(yr,mo)
            if not rows: raise RuntimeError(f'No data in database for {bln} {yr}.')
        label=f"{'DB' if src=='database' else 'Cache'} {bln} {yr}"
        self._generate_and_show(rows,label)

    def _do_all(self):
        self._do_settime()
        self._do_pull()
        self._do_report()

    def _check_update(self):
        try:
            from updater import get_latest_release, is_newer, download_and_replace
        except ImportError:
            messagebox.showinfo("Update","Updater module not found."); return
        self._log("Checking for updates on GitHub...")
        def _run():
            info=get_latest_release()
            if not info:
                self.after(0,lambda:self._log("⚠ Cannot check for updates (no internet or no releases yet)."))
                return
            latest=info['version']; dl_url=info['download_url']; body=info['body'] or ''
            if not is_newer(latest,APP_VERSION):
                self.after(0,lambda:self._log(f'✓ Already up to date (v{APP_VERSION})'))
                self.after(0,lambda:messagebox.showinfo("Update",f"Already up to date!\nCurrent: v{APP_VERSION}"))
                return
            msg=f"New version available: {latest}\n\nChangelog:\n{body[:400]}\n\nDownload & update now?"
            do_it=[False]
            def _ask(): do_it[0]=messagebox.askyesno("Update Available",msg)
            self.after(0,_ask)
            self.after(800,lambda:self._do_download(dl_url,latest) if do_it[0] else None)
        threading.Thread(target=_run,daemon=True).start()

    def _do_download(self,url,version):
        from updater import download_and_replace
        self._log(f"Downloading v{version} ...")
        dlg=tk.Toplevel(self); dlg.title("Downloading"); dlg.resizable(False,False); dlg.grab_set()
        tk.Label(dlg,text=f"Downloading ZKTeco Utility {version}",font=('Segoe UI',10)).pack(padx=20,pady=(14,4))
        pbar=ttk.Progressbar(dlg,length=320,mode='determinate'); pbar.pack(padx=20,pady=6)
        pct_lbl=tk.Label(dlg,text="0%",font=('Segoe UI',9)); pct_lbl.pack(pady=(0,14))
        def on_progress(pct):
            self.after(0,lambda:pbar.configure(value=pct))
            self.after(0,lambda:pct_lbl.configure(text=f"{pct}%"))
        def on_done():
            self.after(0,dlg.destroy)
            self._log(f'✓ Update {version} downloaded — restarting...')
            self.after(0,lambda:messagebox.showinfo("Update Done",f"Updated to {version}!\nApp will restart."))
            self.after(1500,self.destroy)
        def on_error(msg):
            self.after(0,dlg.destroy)
            self.after(0,lambda:messagebox.showerror("Error",f"Download failed:\n{msg}"))
        download_and_replace(url,on_progress=on_progress,on_done=on_done,on_error=on_error)


if __name__=='__main__':
    app=App()
    app.mainloop()
