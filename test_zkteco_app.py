#!/usr/bin/env python3
"""
Unit tests for ZKTeco eFace10 Utility v4.3
Run: python -m pytest test_zkteco_app.py -v
  or: python test_zkteco_app.py
"""

import ast
import csv
import io
import json
import os
import sqlite3
import sys
import tempfile
import unittest
from datetime import datetime, date, timedelta
from unittest.mock import MagicMock, patch, PropertyMock

# ── Mock tkinter before importing app ─────────────────────────────────────────
for mod in ['tkinter', 'tkinter.ttk', 'tkinter.filedialog',
            'tkinter.messagebox', 'tkinter.scrolledtext']:
    sys.modules[mod] = MagicMock()

# ── Import app module ─────────────────────────────────────────────────────────
import importlib.util
_spec = importlib.util.spec_from_file_location(
    "zkteco_app",
    os.path.join(os.path.dirname(__file__), "zkteco_app.py")
)
app = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(app)


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def _make_rows(n=10, base_date=None, uid_start=1, names=None):
    """Generate synthetic attendance rows."""
    if base_date is None:
        base_date = date(2026, 5, 1)
    if names is None:
        names = {1: "NICHOLAS", 2: "SERLI", 3: "TIA"}
    rows = []
    for i in range(n):
        d = base_date + timedelta(days=i % 20)
        uid = (i % len(names)) + uid_start
        name = names.get(uid, f"UID:{uid}")
        # 3 taps per day (simulate redundant scans)
        for hour in [8, 9, 16]:
            rows.append({
                'uid': uid, 'nama': name,
                'timestamp': datetime(d.year, d.month, d.day, hour, i % 60),
                'punch': 0 if hour < 12 else 1,
            })
    return rows


def _tmp_db():
    """Return path to a fresh temp database."""
    fd, path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    return path


# ═════════════════════════════════════════════════════════════════════════════
# 1. CONFIG
# ═════════════════════════════════════════════════════════════════════════════

class TestConfig(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.cfg_path = os.path.join(self.tmpdir, 'config.json')

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_load_config_returns_defaults_when_no_file(self):
        """load_config() returns DEFAULT_CONFIG when file missing."""
        with patch.object(app, 'CONFIG_FILE', self.cfg_path):
            cfg = app.load_config()
        self.assertEqual(cfg['ip'], '10.10.11.55')
        self.assertEqual(cfg['port'], '8088')
        self.assertIn('user_map', cfg)

    def test_save_and_reload_config(self):
        """save_config() persists values that load_config() reads back."""
        with patch.object(app, 'CONFIG_FILE', self.cfg_path):
            cfg = app.load_config()
            cfg['ip'] = '192.168.1.99'
            cfg['lang'] = 'id'
            app.save_config(cfg)
            reloaded = app.load_config()
        self.assertEqual(reloaded['ip'], '192.168.1.99')
        self.assertEqual(reloaded['lang'], 'id')

    def test_load_config_merges_missing_keys(self):
        """load_config() adds missing default keys to existing config."""
        partial = {'ip': '1.2.3.4', 'port': '9999'}
        with open(self.cfg_path, 'w') as f:
            json.dump(partial, f)
        with patch.object(app, 'CONFIG_FILE', self.cfg_path):
            cfg = app.load_config()
        self.assertIn('jam_masuk', cfg)
        self.assertIn('user_map', cfg)
        self.assertEqual(cfg['ip'], '1.2.3.4')

    def test_load_config_handles_corrupt_json(self):
        """load_config() falls back to defaults on corrupt JSON."""
        with open(self.cfg_path, 'w') as f:
            f.write("{corrupt json{{")
        with patch.object(app, 'CONFIG_FILE', self.cfg_path):
            cfg = app.load_config()
        self.assertEqual(cfg['ip'], app.DEFAULT_CONFIG['ip'])


# ═════════════════════════════════════════════════════════════════════════════
# 2. DATABASE
# ═════════════════════════════════════════════════════════════════════════════

class TestDatabase(unittest.TestCase):

    def setUp(self):
        self.db = _tmp_db()
        with patch.object(app, 'DB_FILE', self.db):
            app.init_db()

    def tearDown(self):
        os.unlink(self.db)

    # ── init ──────────────────────────────────────────────────────────────────
    def test_init_db_creates_tables(self):
        conn = sqlite3.connect(self.db)
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        conn.close()
        self.assertIn('attendance', tables)
        self.assertIn('users', tables)
        self.assertIn('pull_sessions', tables)
        self.assertIn('excel_snapshots', tables)

    def test_init_db_idempotent(self):
        """Calling init_db() twice does not raise or duplicate tables."""
        with patch.object(app, 'DB_FILE', self.db):
            app.init_db()   # second call
        conn = sqlite3.connect(self.db)
        cnt = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='attendance'"
        ).fetchone()[0]
        conn.close()
        self.assertEqual(cnt, 1)

    # ── attendance insert ─────────────────────────────────────────────────────
    def test_insert_attendance_basic(self):
        rows = _make_rows(5)
        with patch.object(app, 'DB_FILE', self.db):
            inserted = app.db_insert_attendance(rows)
        self.assertGreater(inserted, 0)

    def test_insert_attendance_deduplicates(self):
        """Inserting same rows twice: second insert adds 0 new records."""
        rows = _make_rows(5)
        with patch.object(app, 'DB_FILE', self.db):
            first  = app.db_insert_attendance(rows)
            second = app.db_insert_attendance(rows)
        self.assertGreater(first, 0)
        self.assertEqual(second, 0)

    def test_insert_attendance_filters_year_2000(self):
        """Records with timestamp year <= 2000 are inserted but filtered on query."""
        stale = [{
            'uid': 1, 'nama': 'TEST',
            'timestamp': datetime(2000, 1, 1, 8, 0),
            'punch': 0,
        }]
        with patch.object(app, 'DB_FILE', self.db):
            app.db_insert_attendance(stale)
            count = app.db_count()
        self.assertEqual(count, 0)   # db_count filters year > 2000

    # ── attendance query ──────────────────────────────────────────────────────
    def test_query_attendance_all(self):
        rows = _make_rows(6)
        with patch.object(app, 'DB_FILE', self.db):
            app.db_insert_attendance(rows)
            result = app.db_query_attendance()
        self.assertGreater(len(result), 0)
        # every result has required keys
        for r in result:
            self.assertIn('uid', r)
            self.assertIn('nama', r)
            self.assertIn('timestamp', r)

    def test_query_attendance_filter_month(self):
        rows = _make_rows(3, base_date=date(2026, 4, 1))
        rows += _make_rows(3, base_date=date(2026, 5, 1))
        with patch.object(app, 'DB_FILE', self.db):
            app.db_insert_attendance(rows)
            april = app.db_query_attendance(year=2026, month=4)
            may   = app.db_query_attendance(year=2026, month=5)
        for r in april:
            self.assertEqual(r['timestamp'].month, 4)
        for r in may:
            self.assertEqual(r['timestamp'].month, 5)

    def test_db_count(self):
        rows = _make_rows(4)
        with patch.object(app, 'DB_FILE', self.db):
            app.db_insert_attendance(rows)
            n = app.db_count()
        self.assertGreater(n, 0)

    # ── users ─────────────────────────────────────────────────────────────────
    def test_upsert_users(self):
        users = [{'uid': 1, 'nama': 'NICHOLAS', 'card_id': '123456'},
                 {'uid': 2, 'nama': 'SERLI',    'card_id': ''}]
        with patch.object(app, 'DB_FILE', self.db):
            app.db_upsert_users(users)
        conn = sqlite3.connect(self.db)
        rows = conn.execute("SELECT uid,nama FROM users ORDER BY uid").fetchall()
        conn.close()
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0][1], 'NICHOLAS')

    def test_upsert_users_updates_existing(self):
        """Upserting same UID with new name updates the record."""
        with patch.object(app, 'DB_FILE', self.db):
            app.db_upsert_users([{'uid': 1, 'nama': 'OLD', 'card_id': ''}])
            app.db_upsert_users([{'uid': 1, 'nama': 'NEW', 'card_id': ''}])
        conn = sqlite3.connect(self.db)
        name = conn.execute("SELECT nama FROM users WHERE uid=1").fetchone()[0]
        conn.close()
        self.assertEqual(name, 'NEW')

    # ── pull sessions ─────────────────────────────────────────────────────────
    def test_add_and_get_pull_sessions(self):
        with patch.object(app, 'DB_FILE', self.db):
            sid = app.db_add_pull_session(100, 50, '10.10.11.55', 'test')
            sessions = app.db_get_pull_sessions()
        self.assertGreater(sid, 0)
        self.assertTrue(any(s[0] == sid for s in sessions))

    def test_delete_pull_session(self):
        with patch.object(app, 'DB_FILE', self.db):
            sid = app.db_add_pull_session(10, 5, '10.10.11.55')
            app.db_delete_pull_session(sid)
            sessions = app.db_get_pull_sessions()
        self.assertFalse(any(s[0] == sid for s in sessions))

    # ── excel snapshots ───────────────────────────────────────────────────────
    def test_save_and_load_snapshot(self):
        payload = b'fake_excel_bytes_here'
        with patch.object(app, 'DB_FILE', self.db):
            snap_id = app.db_save_excel_snapshot(0, 'Test Report', 2026, 5, payload)
            result  = app.db_load_excel_snapshot(snap_id)
        self.assertIsNotNone(result)
        self.assertEqual(result[0], payload)
        self.assertEqual(result[1], 'Test Report')

    def test_delete_snapshot(self):
        with patch.object(app, 'DB_FILE', self.db):
            snap_id = app.db_save_excel_snapshot(0, 'To delete', 2026, None, b'data')
            app.db_delete_excel_snapshot(snap_id)
            result = app.db_load_excel_snapshot(snap_id)
        self.assertIsNone(result)

    def test_get_excel_snapshots_returns_list(self):
        with patch.object(app, 'DB_FILE', self.db):
            app.db_save_excel_snapshot(0, 'Report A', 2026, 4, b'aaa')
            app.db_save_excel_snapshot(0, 'Report B', 2026, 5, b'bbb')
            snaps = app.db_get_excel_snapshots()
        self.assertEqual(len(snaps), 2)

    def test_delete_session_cascades_snapshots(self):
        """Deleting a pull session also removes its snapshots."""
        with patch.object(app, 'DB_FILE', self.db):
            sid = app.db_add_pull_session(10, 5, '10.10.11.55')
            snap_id = app.db_save_excel_snapshot(sid, 'Report', 2026, None, b'data')
            app.db_delete_pull_session(sid)
            result = app.db_load_excel_snapshot(snap_id)
        self.assertIsNone(result)


# ═════════════════════════════════════════════════════════════════════════════
# 3. DEDUPLICATION LOGIC (_DF / _Row)
# ═════════════════════════════════════════════════════════════════════════════

class TestDeduplication(unittest.TestCase):
    """Test that earliest tap = check-in, latest tap = check-out."""

    def _run_dedup(self, taps_by_person):
        """
        taps_by_person: dict {name: [datetime, ...]}
        Returns dict {name: _Row}
        """
        rows = []
        uid_map = {name: i+1 for i,name in enumerate(taps_by_person)}
        for name, taps in taps_by_person.items():
            uid = uid_map[name]
            for ts in taps:
                rows.append({'uid': uid, 'nama': name, 'timestamp': ts, 'punch': 0})

        cfg = dict(app.DEFAULT_CONFIG)
        cfg['user_map'] = {str(v): k for k, v in uid_map.items()}
        data = app.generate_excel_bytes(rows, cfg)
        # We test via parse_excel_for_preview (Log Detail sheet = last sheet)
        headers, preview = app.parse_excel_for_preview(data, -1)
        return headers, preview

    def test_single_tap_no_checkout(self):
        """One tap per day = check-in only, check-out = '-'."""
        taps = [datetime(2026, 5, 1, 8, 0)]
        hdrs, rows = self._run_dedup({'NICHOLAS': taps})
        # find check-out column
        if 'Check-out' in hdrs:
            col = hdrs.index('Check-out')
        elif 'Jam Keluar' in hdrs:
            col = hdrs.index('Jam Keluar')
        else:
            self.skipTest("Check-out column not found in log detail")
        self.assertEqual(rows[0][col], '-')

    def test_multiple_taps_earliest_is_checkin(self):
        """With 5 taps: earliest = check-in."""
        taps = [
            datetime(2026, 5, 1, 7, 30),
            datetime(2026, 5, 1, 8,  0),
            datetime(2026, 5, 1, 12, 0),
            datetime(2026, 5, 1, 16, 0),
            datetime(2026, 5, 1, 16, 35),
        ]
        hdrs, rows = self._run_dedup({'NICHOLAS': taps})
        if 'Check-in' in hdrs:
            col = hdrs.index('Check-in')
        elif 'Jam Masuk' in hdrs:
            col = hdrs.index('Jam Masuk')
        else:
            self.skipTest("Check-in column not found")
        self.assertEqual(rows[0][col], '07:30')

    def test_multiple_taps_latest_is_checkout(self):
        """With 5 taps: latest = check-out."""
        taps = [
            datetime(2026, 5, 1, 7, 30),
            datetime(2026, 5, 1, 8,  0),
            datetime(2026, 5, 1, 12, 0),
            datetime(2026, 5, 1, 16, 0),
            datetime(2026, 5, 1, 16, 35),
        ]
        hdrs, rows = self._run_dedup({'NICHOLAS': taps})
        if 'Check-out' in hdrs:
            col = hdrs.index('Check-out')
        elif 'Jam Keluar' in hdrs:
            col = hdrs.index('Jam Keluar')
        else:
            self.skipTest("Check-out column not found")
        self.assertEqual(rows[0][col], '16:35')

    def test_tap_total_shows_raw_count(self):
        """Taps column shows actual raw tap count from machine."""
        taps = [datetime(2026, 5, 1, h, 0) for h in [8, 9, 10, 16, 16]]
        hdrs, rows = self._run_dedup({'NICHOLAS': taps})
        if 'Taps' in hdrs:
            col = hdrs.index('Taps')
        elif 'Tap' in hdrs:
            col = hdrs.index('Tap')
        else:
            self.skipTest("Taps column not found")
        self.assertEqual(rows[0][col], '5x')


# ═════════════════════════════════════════════════════════════════════════════
# 4. EXCEL GENERATION
# ═════════════════════════════════════════════════════════════════════════════

class TestExcelGeneration(unittest.TestCase):

    def setUp(self):
        self.cfg = dict(app.DEFAULT_CONFIG)
        self.cfg['user_map'] = {'1': 'NICHOLAS', '2': 'SERLI', '3': 'TIA'}
        self.rows = _make_rows(15, base_date=date(2026, 5, 1))

    def test_returns_bytes(self):
        result = app.generate_excel_bytes(self.rows, self.cfg)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_output_is_valid_xlsx(self):
        """Output bytes can be parsed by openpyxl."""
        from openpyxl import load_workbook
        data = app.generate_excel_bytes(self.rows, self.cfg)
        wb = load_workbook(io.BytesIO(data))
        self.assertGreater(len(wb.sheetnames), 0)
        wb.close()

    def test_contains_recap_sheet(self):
        from openpyxl import load_workbook
        data = app.generate_excel_bytes(self.rows, self.cfg)
        wb = load_workbook(io.BytesIO(data))
        self.assertTrue(any('Recap' in s or 'Rekap' in s for s in wb.sheetnames))
        wb.close()

    def test_contains_log_detail_sheet(self):
        from openpyxl import load_workbook
        data = app.generate_excel_bytes(self.rows, self.cfg)
        wb = load_workbook(io.BytesIO(data))
        self.assertIn('Log Detail', wb.sheetnames)
        wb.close()

    def test_raises_on_empty_rows(self):
        with self.assertRaises(RuntimeError):
            app.generate_excel_bytes([], self.cfg)

    def test_raises_on_all_year_2000(self):
        stale = [{'uid':1,'nama':'X','timestamp':datetime(2000,1,1,8,0),'punch':0}]
        with self.assertRaises(RuntimeError):
            app.generate_excel_bytes(stale, self.cfg)

    def test_month_filter_in_sheet_name(self):
        """Each calendar month in data gets its own sheet."""
        from openpyxl import load_workbook
        rows_apr = _make_rows(5, base_date=date(2026, 4, 1))
        rows_may = _make_rows(5, base_date=date(2026, 5, 1))
        data = app.generate_excel_bytes(rows_apr + rows_may, self.cfg)
        wb = load_workbook(io.BytesIO(data))
        names = wb.sheetnames
        wb.close()
        # Should have sheets for both Apr2026 and May2026
        self.assertTrue(any('Apr' in s or 'Apr' in s for s in names))
        self.assertTrue(any('May' in s or 'Mei' in s for s in names))

    def test_english_headers_when_lang_en(self):
        self.cfg['lang'] = 'en'
        data = app.generate_excel_bytes(self.rows, self.cfg)
        hdrs, _ = app.parse_excel_for_preview(data, -1)  # Log Detail
        self.assertTrue(any('Name' in h or 'Check' in h for h in hdrs))

    def test_indonesian_headers_when_lang_id(self):
        self.cfg['lang'] = 'id'
        data = app.generate_excel_bytes(self.rows, self.cfg)
        hdrs, _ = app.parse_excel_for_preview(data, -1)
        self.assertTrue(any('Nama' in h or 'Masuk' in h for h in hdrs))

    def test_no_file_created_on_disk(self):
        """generate_excel_bytes must not write any file to disk."""
        before = set(os.listdir('.'))
        app.generate_excel_bytes(self.rows, self.cfg)
        after = set(os.listdir('.'))
        new_files = {f for f in (after - before) if f.endswith('.xlsx')}
        self.assertEqual(new_files, set())


# ═════════════════════════════════════════════════════════════════════════════
# 5. LATE / OVERTIME CALCULATION
# ═════════════════════════════════════════════════════════════════════════════

class TestLateOvertimeCalc(unittest.TestCase):

    def _gen_one_day(self, check_in, check_out, jam_masuk='08:00',
                     jam_keluar='16:00', toleransi=15):
        cfg = dict(app.DEFAULT_CONFIG)
        cfg['jam_masuk']  = jam_masuk
        cfg['jam_keluar'] = jam_keluar
        cfg['toleransi']  = toleransi
        cfg['lang']       = 'en'
        cfg['user_map']   = {'1': 'NICHOLAS'}
        rows = [
            {'uid':1,'nama':'NICHOLAS',
             'timestamp': datetime(2026,5,4,*map(int,check_in.split(':'))), 'punch':0},
            {'uid':1,'nama':'NICHOLAS',
             'timestamp': datetime(2026,5,4,*map(int,check_out.split(':'))), 'punch':1},
        ]
        data = app.generate_excel_bytes(rows, cfg)
        hdrs, preview = app.parse_excel_for_preview(data, -1)
        if not preview: return {}
        row = dict(zip(hdrs, preview[0]))
        return row

    def test_on_time_no_late(self):
        row = self._gen_one_day('08:00', '16:00')
        late = row.get('Late') or row.get('Terlambat') or '-'
        self.assertEqual(late, '-')

    def test_within_tolerance_no_late(self):
        """08:14 with 15-min tolerance = not late."""
        row = self._gen_one_day('08:14', '16:00', toleransi=15)
        late = row.get('Late') or row.get('Terlambat') or '-'
        self.assertEqual(late, '-')

    def test_over_tolerance_is_late(self):
        """08:20 with 15-min tolerance = 20 min late."""
        row = self._gen_one_day('08:20', '16:00', toleransi=15)
        late = row.get('Late') or row.get('Terlambat') or ''
        self.assertIn('20', str(late))

    def test_no_overtime_on_time(self):
        row = self._gen_one_day('08:00', '16:00')
        ot = row.get('OT') or row.get('Lembur') or '-'
        self.assertEqual(ot, '-')

    def test_overtime_detected(self):
        """Check-out at 18:00 = 120 min OT (std 16:00)."""
        row = self._gen_one_day('08:00', '18:00')
        ot = row.get('OT') or row.get('Lembur') or ''
        self.assertIn('120', str(ot))

    def test_early_checkin_not_counted_as_late(self):
        """Arriving at 07:30 should not be marked late."""
        row = self._gen_one_day('07:30', '16:00')
        late = row.get('Late') or row.get('Terlambat') or '-'
        self.assertEqual(late, '-')


# ═════════════════════════════════════════════════════════════════════════════
# 6. PARSE EXCEL FOR PREVIEW
# ═════════════════════════════════════════════════════════════════════════════

class TestParseExcelForPreview(unittest.TestCase):

    def setUp(self):
        cfg = dict(app.DEFAULT_CONFIG)
        cfg['user_map'] = {'1': 'NICHOLAS', '2': 'SERLI'}
        cfg['lang'] = 'en'
        rows = _make_rows(6, base_date=date(2026, 5, 1))
        self.data = app.generate_excel_bytes(rows, cfg)

    def test_returns_headers_and_rows(self):
        headers, rows = app.parse_excel_for_preview(self.data, 0)
        self.assertIsInstance(headers, list)
        self.assertIsInstance(rows, list)
        self.assertGreater(len(headers), 0)

    def test_all_rows_have_same_col_count(self):
        headers, rows = app.parse_excel_for_preview(self.data, 0)
        for r in rows:
            self.assertEqual(len(r), len(headers),
                             f"Row has {len(r)} cols, expected {len(headers)}")

    def test_can_parse_all_sheets(self):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(self.data), read_only=True)
        n_sheets = len(wb.sheetnames)
        wb.close()
        for i in range(n_sheets):
            headers, rows = app.parse_excel_for_preview(self.data, i)
            self.assertIsInstance(headers, list)

    def test_handles_invalid_bytes(self):
        headers, rows = app.parse_excel_for_preview(b'not_excel_data', 0)
        self.assertIsInstance(rows, list)
        # should not raise, just return error info


# ═════════════════════════════════════════════════════════════════════════════
# 7. OPEN_PATH (cross-platform)
# ═════════════════════════════════════════════════════════════════════════════

class TestOpenPath(unittest.TestCase):

    def test_windows_uses_startfile(self):
        # os.startfile only exists on Windows; use create=True for Linux CI
        with patch.object(sys, 'platform', 'win32'), \
             patch('os.startfile', create=True) as mock_sf:
            app._open_path('/some/path')
            mock_sf.assert_called_once_with('/some/path')

    def test_macos_uses_open(self):
        with patch.object(sys, 'platform', 'darwin'), \
             patch('subprocess.Popen') as mock_popen:
            app._open_path('/some/path')
            mock_popen.assert_called_once_with(['open', '/some/path'])

    def test_linux_uses_xdg_open(self):
        with patch.object(sys, 'platform', 'linux'), \
             patch('subprocess.Popen') as mock_popen:
            app._open_path('/some/path')
            mock_popen.assert_called_once_with(['xdg-open', '/some/path'])

    def test_does_not_raise_on_error(self):
        with patch.object(sys, 'platform', 'win32'), \
             patch('os.startfile', create=True, side_effect=OSError("fail")):
            app._open_path('/bad/path')  # must not propagate


# ═════════════════════════════════════════════════════════════════════════════
# 8. SYNTAX & IMPORTS
# ═════════════════════════════════════════════════════════════════════════════

class TestAppSyntax(unittest.TestCase):

    def test_app_parses_without_syntax_errors(self):
        src = open(os.path.join(os.path.dirname(__file__), 'zkteco_app.py'),
                   encoding='utf-8').read()
        try:
            ast.parse(src)
        except SyntaxError as e:
            self.fail(f"Syntax error in zkteco_app.py: {e}")

    def test_i18n_parses_without_syntax_errors(self):
        path = os.path.join(os.path.dirname(__file__), 'i18n.py')
        if not os.path.exists(path):
            self.skipTest("i18n.py not found")
        src = open(path, encoding='utf-8').read()
        try:
            ast.parse(src)
        except SyntaxError as e:
            self.fail(f"Syntax error in i18n.py: {e}")

    def test_app_version_string_is_valid(self):
        parts = app.APP_VERSION.split('.')
        self.assertEqual(len(parts), 3)
        for p in parts:
            self.assertTrue(p.isdigit(), f"Version part '{p}' is not numeric")

    def test_default_config_has_required_keys(self):
        required = ['ip', 'port', 'jam_masuk', 'jam_keluar',
                    'toleransi', 'lang', 'auto_backup', 'user_map']
        for key in required:
            self.assertIn(key, app.DEFAULT_CONFIG, f"Missing key: {key}")

    def test_db_file_and_config_file_are_absolute_paths(self):
        self.assertTrue(os.path.isabs(app.DB_FILE))
        self.assertTrue(os.path.isabs(app.CONFIG_FILE))


# ═════════════════════════════════════════════════════════════════════════════
# 9. INTEGRATION — full pull → generate → snapshot → load flow
# ═════════════════════════════════════════════════════════════════════════════

class TestIntegration(unittest.TestCase):

    def setUp(self):
        self.db = _tmp_db()
        self.tmpdir = tempfile.mkdtemp()
        self.cfg_path = os.path.join(self.tmpdir, 'config.json')
        with patch.object(app, 'DB_FILE', self.db):
            app.init_db()

    def tearDown(self):
        import shutil
        os.unlink(self.db)
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_full_pipeline(self):
        """Simulate: pull → insert → generate → save snapshot → load snapshot."""
        # 1. Simulate pulling from device
        pulled_rows = _make_rows(20, base_date=date(2026, 5, 1))

        with patch.object(app, 'DB_FILE', self.db):
            # 2. Insert to DB
            new = app.db_insert_attendance(pulled_rows)
            self.assertGreater(new, 0)

            # 3. Record pull session
            sid = app.db_add_pull_session(len(pulled_rows), new, '10.10.11.55')
            self.assertGreater(sid, 0)

            # 4. Generate Excel
            cfg = dict(app.DEFAULT_CONFIG)
            cfg['user_map'] = {'1':'NICHOLAS','2':'SERLI','3':'TIA'}
            data = app.generate_excel_bytes(pulled_rows, cfg)
            self.assertIsInstance(data, bytes)
            self.assertGreater(len(data), 1000)

            # 5. Save snapshot to DB
            snap_id = app.db_save_excel_snapshot(sid, 'Integration Test', 2026, 5, data)
            self.assertGreater(snap_id, 0)

            # 6. Load snapshot back
            loaded = app.db_load_excel_snapshot(snap_id)
            self.assertIsNotNone(loaded)
            self.assertEqual(loaded[0], data)   # bytes match exactly
            self.assertEqual(loaded[1], 'Integration Test')

            # 7. Preview loaded data
            hdrs, rows = app.parse_excel_for_preview(loaded[0], 0)
            self.assertGreater(len(hdrs), 0)
            self.assertGreater(len(rows), 0)

            # 8. History reflects session & snapshot
            sessions = app.db_get_pull_sessions()
            snaps    = app.db_get_excel_snapshots()
            self.assertTrue(any(s[0]==sid for s in sessions))
            self.assertTrue(any(s[0]==snap_id for s in snaps))

            # 9. Delete session cascades snapshot
            app.db_delete_pull_session(sid)
            self.assertIsNone(app.db_load_excel_snapshot(snap_id))

    def test_multi_month_data_generates_correct_sheets(self):
        """Data spanning 3 months → 3 calendar sheets + Recap + Log Detail."""
        from openpyxl import load_workbook
        rows = []
        for mo in [3, 4, 5]:
            rows += _make_rows(5, base_date=date(2026, mo, 1))
        cfg = dict(app.DEFAULT_CONFIG)
        cfg['user_map'] = {'1':'NICHOLAS','2':'SERLI','3':'TIA'}
        data = app.generate_excel_bytes(rows, cfg)
        wb = load_workbook(io.BytesIO(data), read_only=True)
        names = wb.sheetnames
        wb.close()
        # At least 3 calendar sheets + recap + log
        self.assertGreaterEqual(len(names), 5)


# ═════════════════════════════════════════════════════════════════════════════
# RUNNER
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    loader = unittest.TestLoader()
    suite  = unittest.TestSuite()

    test_classes = [
        TestConfig,
        TestDatabase,
        TestDeduplication,
        TestExcelGeneration,
        TestLateOvertimeCalc,
        TestParseExcelForPreview,
        TestOpenPath,
        TestAppSyntax,
        TestIntegration,
    ]

    for cls in test_classes:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
    result = runner.run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
